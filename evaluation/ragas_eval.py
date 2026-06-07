"""RAGAS 지표 계산 — 기존 평가 결과 JSON에 RAGAS 점수를 추가한다.

evaluate.py 로 평가를 먼저 완료한 뒤, 이 스크립트로 RAGAS 지표를 별도 계산한다.
계산 결과는 원본 JSON 을 덮어쓰고, 같은 이름의 xlsx 도 업데이트한다.

실행:
    python evaluation/ragas_eval.py                          # 최신 결과 자동 선택
    python evaluation/ragas_eval.py evaluation_results_*.json  # 파일 직접 지정

지표:
    faithfulness      : 답변이 컨텍스트에 근거한 정도 (0~1)
    answer_relevancy  : 답변이 질문과 관련된 정도 (0~1)
    context_precision : 검색된 컨텍스트 중 관련 있는 비율 (0~1)
    context_recall    : ground_truth 를 컨텍스트가 커버하는 정도 (0~1)
"""
import os
import sys
import json
import glob
import warnings
warnings.filterwarnings("ignore")

from dotenv import load_dotenv

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR, ".env"))

RESULTS_DIR = os.path.join(BASE_DIR, "evaluation", "results")
_RAGAS_METRICS = ("faithfulness", "answer_relevancy", "context_precision", "context_recall")


def _pick_json() -> str:
    if len(sys.argv) > 1:
        path = sys.argv[1]
        if not os.path.isabs(path):
            path = os.path.join(RESULTS_DIR, path)
        return path
    files = sorted(glob.glob(os.path.join(RESULTS_DIR, "evaluation_results_*.json")))
    if not files:
        raise FileNotFoundError(f"결과 JSON 없음: {RESULTS_DIR}")
    return files[-1]


def compute_ragas(rows: list) -> None:
    """rows 에 RAGAS 지표를 in-place 로 추가."""
    try:
        from datasets import Dataset as HFDataset
        from ragas import evaluate as ragas_evaluate
        from ragas.metrics import (
            faithfulness, answer_relevancy,
            context_precision, context_recall,
        )
        from ragas.llms import LangchainLLMWrapper
        from ragas.embeddings import LangchainEmbeddingsWrapper
        from langchain_upstage import ChatUpstage, UpstageEmbeddings
    except ImportError as e:
        sys.exit(f"[오류] 패키지 누락: {e}\n  pip install ragas langchain-upstage")

    api_key = os.getenv("UPSTAGE_API_KEY")
    if not api_key:
        sys.exit("[오류] UPSTAGE_API_KEY 가 .env 에 없음")

    llm = LangchainLLMWrapper(ChatUpstage(api_key=api_key, model="solar-pro2"))
    emb = LangchainEmbeddingsWrapper(UpstageEmbeddings(api_key=api_key, model="embedding-query"))
    faithfulness.llm = llm
    answer_relevancy.llm = llm
    answer_relevancy.embeddings = emb
    context_precision.llm = llm
    context_recall.llm = llm
    metrics = [faithfulness, answer_relevancy, context_precision, context_recall]

    for prefix in ("vector", "hybrid"):
        answer_key = f"{prefix}_answer"
        ctx_key = f"{prefix}_contexts"
        valid = [(i, r) for i, r in enumerate(rows)
                 if r.get(answer_key, "").strip() and r.get(ctx_key)]
        if not valid:
            print(f"[RAGAS] {prefix}: 유효 행 없음, 건너뜀")
            continue

        print(f"[RAGAS] {prefix}: {len(valid)}개 계산 중...")
        data = {
            "user_input":         [r["question"]              for _, r in valid],
            "response":           [r[answer_key]               for _, r in valid],
            "retrieved_contexts": [r.get(ctx_key, [])          for _, r in valid],
            "reference":          [r.get("ground_truth", "")   for _, r in valid],
        }
        try:
            dataset = HFDataset.from_dict(data)
            result = ragas_evaluate(dataset, metrics=metrics)
            result_df = result.to_pandas()
            for df_idx, (_, row) in enumerate(valid):
                if df_idx < len(result_df):
                    for m in _RAGAS_METRICS:
                        val = result_df.iloc[df_idx].get(m)
                        row[f"{prefix}_ragas_{m}"] = (
                            round(float(val), 4)
                            if val is not None and val == val else None
                        )
            print(f"[RAGAS] {prefix}: 완료")
        except Exception as e:
            print(f"[RAGAS] {prefix} 계산 실패: {e}")


def _print_summary(rows: list) -> None:
    for prefix in ("vector", "hybrid"):
        print(f"\n=== {prefix.upper()} ===")
        for m in _RAGAS_METRICS:
            vals = [r[f"{prefix}_ragas_{m}"] for r in rows
                    if r.get(f"{prefix}_ragas_{m}") is not None]
            avg = sum(vals) / len(vals) if vals else None
            print(f"  {m:25s}: {avg:.4f}" if avg is not None else f"  {m:25s}: -")


def _update_xlsx(xlsx_path: str, rows: list) -> None:
    """RAGAS 결과를 xlsx 에 반영한다.
    - 기존 xlsx 가 있으면 ③ 세부지표 시트에 RAGAS 컬럼을 추가/업데이트.
    - 없으면 RAGAS 결과만 담은 새 xlsx 를 생성.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[xlsx] openpyxl 없음 — xlsx 건너뜀")
        return

    BLUE_FILL = PatternFill("solid", fgColor="1F497D")
    H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    N_FONT = Font(name="Arial", size=9)
    AL_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_TC = Alignment(horizontal="center", vertical="top")
    AL_TL = Alignment(horizontal="left",   vertical="top",   wrap_text=True)

    ragas_headers = [
        "V_faithfulness", "V_answer_relevancy", "V_context_precision", "V_context_recall",
        "H_faithfulness", "H_answer_relevancy", "H_context_precision", "H_context_recall",
    ]
    ragas_keys = [
        "vector_ragas_faithfulness", "vector_ragas_answer_relevancy",
        "vector_ragas_context_precision", "vector_ragas_context_recall",
        "hybrid_ragas_faithfulness", "hybrid_ragas_answer_relevancy",
        "hybrid_ragas_context_precision", "hybrid_ragas_context_recall",
    ]

    sheet_name = "③ 세부지표"

    # ── 기존 xlsx 업데이트 ──
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
        if sheet_name not in wb.sheetnames:
            print(f"[xlsx] '{sheet_name}' 시트 없음 — 건너뜀")
            return
        ws = wb[sheet_name]
        existing = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        id_col = existing.get("번호", 2)

        start_col = ws.max_column + 1
        col_map = {}
        for i, h in enumerate(ragas_headers):
            if h not in existing:
                c = start_col + i
                cell = ws.cell(1, c, value=h)
                cell.font = H_FONT; cell.fill = BLUE_FILL; cell.alignment = AL_C
                ws.column_dimensions[cell.column_letter].width = 14
                col_map[h] = c
            else:
                col_map[h] = existing[h]

        row_by_id = {str(r.get("id", "")): r for r in rows}
        for row_idx in range(2, ws.max_row + 1):
            qid = str(ws.cell(row_idx, id_col).value or "")
            if qid not in row_by_id:
                continue
            r = row_by_id[qid]
            for h, key in zip(ragas_headers, ragas_keys):
                cell = ws.cell(row_idx, col_map[h], value=r.get(key))
                cell.font = N_FONT; cell.alignment = AL_TC

    # ── 새 xlsx 생성 ──
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(sheet_name)
        base = ["번호", "분류", "질문", "정답"]
        all_headers = base + ragas_headers
        for c, h in enumerate(all_headers, 1):
            cell = ws.cell(1, c, value=h)
            cell.font = H_FONT; cell.fill = BLUE_FILL; cell.alignment = AL_C
            ws.column_dimensions[cell.column_letter].width = 14
        ws.row_dimensions[1].height = 28

        for row_idx, r in enumerate(rows, 2):
            vals = [r.get("id",""), r.get("category",""), r.get("question",""), r.get("ground_truth","")]
            vals += [r.get(k) for k in ragas_keys]
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row_idx, c, value=val)
                cell.font = N_FONT
                cell.alignment = AL_TL if c in (3, 4) else AL_TC

    wb.save(xlsx_path)
    print(f"[xlsx] 저장 완료: {xlsx_path}")


def main():
    json_path = _pick_json()
    print(f"[대상] {json_path}")

    with open(json_path, encoding="utf-8") as f:
        rows = json.load(f)
    if isinstance(rows, dict):
        rows = rows.get("results", rows)

    compute_ragas(rows)

    # JSON 저장
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    print(f"\n[JSON 저장] {json_path}")

    # xlsx 업데이트
    xlsx_path = json_path.replace(".json", ".xlsx")
    _update_xlsx(xlsx_path, rows)

    _print_summary(rows)


if __name__ == "__main__":
    main()
